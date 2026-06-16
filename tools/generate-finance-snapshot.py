#!/usr/bin/env python3
"""Read-only parse of Bob's financial workbook → reports/finance-snapshot.json"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from statistics import median

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path("/Users/bobhopp/FINANCIAL/2021-2026 Financial Workbook.xlsx")
OUT_PATH = ROOT / "reports" / "finance-snapshot.json"

EXCLUDE_APPLE = {"Payment", "Credit", "Debit"}
ONE_TIME_INCOME = {"Tax Refund", "Stimulus", "Refinance", "Empower"}
TRANSFER_WITHDRAWALS = {"Apple", "APPLE", "Amazon", "Wells Fargo", ""}

UTILITY_CATS = {
    "Power and Light",
    "Water Utility",
    "Internet",
    "Telephone",
    "Cell Phone",
    "Gas",
    "Helium",
    "Utilities",
    "Flood Insurance",
}

AI_CAT_RE = re.compile(r"^(AI\s*-|Software\s*-\s*ChatGPT|Software\s*-\s*Midjourney|SUB\s*-\s*Grok)", re.I)
AI_TEXT_RE = re.compile(
    r"\b(chatgpt|openai|cursor|grok|runpod|vercel|gemini|genspark|youmind|abacus|midjourney|creative fabrica|kittl)\b",
    re.I,
)

RETRO_OPS_CATS = {
    "3D Printing",
    "Engraving",
    "Web - Cloudflare",
    "Web - Neon.tech",
    "Web - Netlify",
    "Web - Plex",
    "Software - Hazel",
    "Software - Lossless Cut",
    "Storage - Dropbox",
    "SUB - CloudFlare",
}

RETRO_GEAR_CATS = {"Audio Eq", "Lighting Eq", "Video Eq", "Music eq", "Music Eq", "OWI", "Computer", "Inventory"}
RETRO_AMZ_ACCOUNTS = {"inventory", "Inventory", "Audio Eq", "Video Eq", "Lighting Eq", "Music eq", "Music Eq", "Computer", "Office"}

VENDOR_NORMALIZE = {
    "Software - ChatGPT": "ChatGPT",
    "AI - ChatGPT": "ChatGPT",
    "SUB - Grok": "Grok",
    "AI - Grok": "Grok",
    "AI - Cursor": "Cursor",
    "AI - Runpod.io": "Runpod",
    "AI - Vercel": "Vercel",
    "Software - Adobe": "Adobe",
    "SUB - Adobe": "Adobe",
    "Software - KITTL": "KITTL",
    "SUB - Kittle": "KITTL",
    "AI - Creative Fabrica": "Creative Fabrica",
    "SUB - CreativeFabrica": "Creative Fabrica",
    "AI - Genspark": "GenSpark",
    "SUB - GenSpark Ai": "GenSpark",
    "Storage - Dropbox": "Dropbox",
    "SUB - Dropbox": "Dropbox",
    "Web - Cloudflare": "Cloudflare",
    "SUB - CloudFlare": "Cloudflare",
    "Software - iCloud": "iCloud",
    "Storage - iCloud": "iCloud",
    "SUB - TV": "YouTube TV",
    "SUB - YouTube": "YouTube Premium",
}

KNOWN_AI_VENDORS = {
    "ChatGPT",
    "Cursor",
    "Grok",
    "Gemini",
    "GenSpark",
    "Youmind",
    "Abacus",
    "Creative Fabrica",
    "Runpod",
    "Vercel",
    "Midjourney",
    "KITTL",
    "OpenAI",
}


def parse_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:19], fmt).date()
        except ValueError:
            pass
    return None


def money(n: float) -> float:
    return round(float(n), 2)


def health_ai(monthly: float, pct_change: float) -> str:
    if monthly >= 250 or pct_change >= 100:
        return "Problem"
    if monthly >= 150 or pct_change >= 50:
        return "Review"
    if monthly >= 75 or pct_change >= 25:
        return "Watch"
    return "Healthy"


def health_retro(ops_monthly: float) -> str:
    if ops_monthly >= 800:
        return "Problem"
    if ops_monthly >= 500:
        return "Review"
    if ops_monthly >= 350:
        return "Watch"
    return "Healthy"


def health_cashflow(net: float) -> str:
    if net < -500:
        return "Problem"
    if net < 0:
        return "Review"
    if net < 200:
        return "Watch"
    return "Healthy"


def sub_status(vendor: str, monthly: float, flagged: bool) -> str:
    if flagged:
        return "Review"
    if monthly >= 50:
        return "Watch"
    return "Healthy"


def month_keys_through(end: date, count: int = 12) -> list[str]:
    keys: list[str] = []
    y, m = end.year, end.month
    for _ in range(count):
        keys.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(keys))


def trend_series(counter: Counter, keys: list[str]) -> list[dict]:
    months = [
        {
            "month": datetime.strptime(key, "%Y-%m").strftime("%b %y"),
            "amount": money(counter.get(key, 0)),
        }
        for key in keys
    ]
    avg = sum(m["amount"] for m in months) / max(len(months), 1)
    return months, money(avg), money(avg * 12)


def is_retro_apple(a: dict) -> bool:
    is_ai_dev = a["cat"].startswith("AI -") or a["cat"] in {
        "Software - ChatGPT",
        "AI - Cursor",
        "AI - Vercel",
        "AI - Runpod.io",
    }
    return a["cat"] in RETRO_OPS_CATS or is_ai_dev


def is_ai_item(a: dict) -> bool:
    text = f"{a['cat']} {a['desc']} {a['merchant']}"
    return bool(AI_CAT_RE.match(a["cat"]) or AI_TEXT_RE.search(text))


def main() -> None:
    workbook = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not workbook.exists():
        print(f"Workbook not found: {workbook}", file=sys.stderr)
        sys.exit(1)

    mtime = datetime.fromtimestamp(workbook.stat().st_mtime).isoformat()
    wb = load_workbook(workbook, read_only=True, data_only=True)

    apple = []
    for i, row in enumerate(wb["APPLE"].iter_rows(values_only=True)):
        if i == 0:
            continue
        dt, year, desc, merchant, cat, typ, amt = row[:7]
        if amt is None:
            continue
        try:
            amount = float(amt)
        except (TypeError, ValueError):
            continue
        apple.append(
            {
                "date": parse_date(dt),
                "year": int(year) if year else None,
                "desc": str(desc or ""),
                "merchant": str(merchant or ""),
                "cat": str(cat or "Uncategorized"),
                "amount": amount,
            }
        )

    amazon = []
    for i, row in enumerate(wb["AMAZON"].iter_rows(values_only=True)):
        if i == 0:
            continue
        try:
            amount = float(row[4])
        except (TypeError, ValueError):
            continue
        if amount <= 0:
            continue
        amazon.append(
            {
                "date": parse_date(row[0]),
                "year": int(row[1]) if row[1] else None,
                "account": str(row[2] or ""),
                "title": str(row[3] or ""),
                "amount": amount,
            }
        )

    nebat = []
    for i, row in enumerate(wb["NEBAT"].iter_rows(values_only=True)):
        if i == 0:
            continue
        w = float(row[5]) if row[5] not in (None, "") else 0.0
        d = float(row[6]) if row[6] not in (None, "") else 0.0
        nebat.append(
            {
                "date": parse_date(row[0]),
                "year": int(row[1]) if row[1] else None,
                "account": str(row[3] or ""),
                "desc": str(row[4] or ""),
                "withdrawal": w,
                "deposit": d,
                "balance": float(row[8]) if len(row) > 8 and row[8] not in (None, "") else None,
            }
        )

    wb.close()

    spending = [a for a in apple if a["amount"] > 0 and a["cat"] not in EXCLUDE_APPLE]
    current_year = max(a["year"] or 0 for a in spending)
    ytd = [a for a in spending if a["year"] == current_year]
    prior_year = [a for a in spending if a["year"] == current_year - 1]

    apple_lifetime = sum(a["amount"] for a in spending)
    apple_ytd = sum(a["amount"] for a in ytd)
    months_span = max(
        1,
        len({(a["year"], a["date"].month) for a in spending if a["date"] and a["year"]}),
    )
    apple_monthly = apple_lifetime / months_span

    amazon_lifetime = sum(a["amount"] for a in amazon)
    amazon_ytd = sum(a["amount"] for a in amazon if a["year"] == current_year)
    amazon_months = max(1, len({(a["year"], a["date"].month) for a in amazon if a["date"] and a["year"]}))
    amazon_monthly = amazon_lifetime / amazon_months

    # Income
    income_by_acct = Counter()
    income_ytd = 0.0
    for n in nebat:
        if n["deposit"] <= 0:
            continue
        acct = n["account"] or "Other"
        if acct in ONE_TIME_INCOME:
            continue
        if acct == "Open":
            continue
        income_by_acct[acct] += n["deposit"]
        if n["year"] == current_year:
            income_ytd += n["deposit"]

    income_months = max(1, len({(n["year"], n["date"].month) for n in nebat if n["deposit"] > 0 and n["date"]}))
    recurring_income = sum(income_by_acct.values())
    income_monthly = recurring_income / income_months

    # Required bills
    bill_lines = Counter()
    for a in spending:
        if a["cat"] in UTILITY_CATS or a["cat"] == "SUB - TV":
            bill_lines[a["cat"]] += a["amount"]
    for n in nebat:
        if n["withdrawal"] > 0 and n["account"] in {"Home Loan", "Car Insurance"}:
            bill_lines[n["account"]] += n["withdrawal"]
    bills_monthly = sum(bill_lines.values()) / months_span

    # AI
    ai_items = [a for a in spending if is_ai_item(a)]
    ai_lifetime = sum(a["amount"] for a in ai_items)
    ai_ytd = sum(a["amount"] for a in ai_items if a["year"] == current_year)
    ai_prior = sum(a["amount"] for a in ai_items if a["year"] == current_year - 1)
    ytd_months = max(1, len({(a["year"], a["date"].month) for a in ai_items if a["year"] == current_year and a["date"]}))
    ai_monthly_ytd = ai_ytd / ytd_months
    ai_run_rate = ai_monthly_ytd * 12
    ai_pct = (ai_lifetime / apple_lifetime * 100) if apple_lifetime else 0
    pct_change = ((ai_ytd - ai_prior) / ai_prior * 100) if ai_prior > 0 else (100 if ai_ytd > 0 else 0)

    ai_tools = sorted(
        {
            VENDOR_NORMALIZE.get(a["cat"], a["cat"].replace("AI - ", "").replace("Software - ", ""))
            for a in ai_items
            if VENDOR_NORMALIZE.get(a["cat"], a["cat"].replace("AI - ", "").replace("Software - ", ""))
            in KNOWN_AI_VENDORS
            or a["cat"].startswith(("AI -", "SUB - Grok", "Software - ChatGPT", "Software - Midjourney"))
        }
    )

    ai_flags = []
    chatgpt_total = sum(a["amount"] for a in ai_items if "ChatGPT" in a["cat"] or "chatgpt" in a["desc"].lower())
    if sum(1 for c in {a["cat"] for a in ai_items} if "ChatGPT" in c) > 1:
        ai_flags.append("ChatGPT appears in multiple categories — confirm one subscription")
    if any(a["cat"].startswith("SUB - Grok") for a in ai_items) and chatgpt_total > 0:
        ai_flags.append("Grok and ChatGPT both active — possible overlap")
    if len(ai_tools) >= 5:
        ai_flags.append(f"{len(ai_tools)} AI tools billed — trial churn risk")

    data_through = max(
        [a["date"] for a in spending if a["date"]]
        + [a["date"] for a in amazon if a["date"]]
        + [n["date"] for n in nebat if n["date"]],
        default=date.today(),
    )
    trend_end = min(data_through, date.today())
    data_through_iso = data_through.isoformat()
    trend_keys = month_keys_through(trend_end, 12)

    ai_monthly_counter: Counter = Counter()
    for a in ai_items:
        if a["date"]:
            ai_monthly_counter[a["date"].strftime("%Y-%m")] += a["amount"]
    ai_trend_months, ai_trend_avg, ai_trend_annual = trend_series(ai_monthly_counter, trend_keys)

    retro_monthly_counter: Counter = Counter()
    for a in spending:
        if a["date"] and is_retro_apple(a):
            retro_monthly_counter[a["date"].strftime("%Y-%m")] += a["amount"]
    for a in amazon:
        if a["date"] and a["account"] in RETRO_AMZ_ACCOUNTS:
            retro_monthly_counter[a["date"].strftime("%Y-%m")] += a["amount"]
    retro_trend_months, retro_trend_avg, retro_trend_annual = trend_series(retro_monthly_counter, trend_keys)

    # Retroverse
    retro_ops = Counter()
    retro_gear_ytd = 0.0
    for a in spending:
        is_ai_dev = a["cat"].startswith("AI -") or a["cat"] in {"Software - ChatGPT", "AI - Cursor", "AI - Vercel", "AI - Runpod.io"}
        if a["cat"] in RETRO_OPS_CATS or is_ai_dev:
            retro_ops[a["cat"]] += a["amount"]
        if a["cat"] in RETRO_GEAR_CATS:
            if a["year"] == current_year:
                retro_gear_ytd += a["amount"]
    for a in amazon:
        if a["account"] in RETRO_GEAR_CATS or a["account"] in RETRO_AMZ_ACCOUNTS:
            if a["year"] == current_year:
                retro_gear_ytd += a["amount"]
        if a["account"] in RETRO_AMZ_ACCOUNTS and a["year"] == current_year:
            retro_ops[f"Amazon · {a['account']}"] += a["amount"]

    retro_ops_total = sum(retro_ops.values())
    retro_ops_ytd = sum(
        a["amount"]
        for a in spending
        if a["year"] == current_year
        and (
            a["cat"] in RETRO_OPS_CATS
            or a["cat"].startswith("AI -")
            or a["cat"] in {"Software - ChatGPT", "AI - Cursor", "AI - Vercel", "AI - Runpod.io"}
        )
    )
    for a in amazon:
        if a["year"] == current_year and a["account"] in RETRO_AMZ_ACCOUNTS:
            retro_ops_ytd += a["amount"]

    retro_ytd_months = max(
        1,
        len({(a["year"], a["date"].month) for a in spending if a["year"] == current_year and a["date"]}),
    )
    retro_ops_monthly = retro_ops_ytd / retro_ytd_months

    # Subscriptions
    sub_prefixes = ("SUB -", "SUB _", "AI -", "Software -", "Storage -", "Web -", "App -")
    sub_cats = defaultdict(list)
    for a in spending:
        if a["cat"].startswith(sub_prefixes):
            vendor = VENDOR_NORMALIZE.get(a["cat"], a["cat"].replace("SUB - ", "").replace("SUB _ ", ""))
            sub_cats[vendor].append(a)

    hosting_monthly = (
        retro_ops.get("Web - Cloudflare", 0) + retro_ops.get("Web - Neon.tech", 0) + retro_ops.get("Web - Netlify", 0)
    ) / months_span

    sub_rows = []
    active_subs = []
    flagged_vendors = {"ChatGPT", "Grok", "GenSpark", "Creative Fabrica", "KITTL", "Adobe", "Dropbox", "Cloudflare", "iCloud"}
    today = trend_end
    for vendor, txns in sub_cats.items():
        total = sum(t["amount"] for t in txns)
        if len(txns) == 1 and total > 120:
            continue  # one-time software purchase
        months_active = max(1, len({(t["year"], t["date"].month) for t in txns if t["date"]}))
        est_monthly = total / months_active
        note = None
        flagged = vendor in flagged_vendors and vendor in {"ChatGPT", "Grok", "GenSpark", "Creative Fabrica"}
        if vendor == "ChatGPT" and len({t["cat"] for t in txns}) > 1:
            note = "Duplicate workbook tags"
            flagged = True
        dated = [t for t in txns if t["date"]]
        last_charge = max(t["date"] for t in dated) if dated else None
        days_since = (today - last_charge).days if last_charge else 999
        is_active = len(txns) >= 2 or days_since <= 120
        row = {
            "vendor": vendor,
            "monthly": money(est_monthly),
            "lifetime": money(total),
            "status": sub_status(vendor, est_monthly, flagged),
            "note": note,
        }
        sub_rows.append(row)
        if is_active:
            active_subs.append(
                {
                    "vendor": vendor,
                    "monthly": money(est_monthly),
                    "annual": money(est_monthly * 12),
                    "lastCharge": last_charge.isoformat() if last_charge else "",
                    "status": row["status"],
                    "note": note,
                }
            )
    sub_rows.sort(key=lambda r: -r["monthly"])
    active_subs.sort(key=lambda r: -r["monthly"])
    top_subs = sub_rows[:5]
    more = sub_rows[5:]
    more_monthly = sum(r["monthly"] for r in more)

    # Opportunity — last 12 months vs prior 12 months by category
    prior_end = datetime.strptime(trend_keys[0], "%Y-%m").date().replace(day=1) - timedelta(days=1)
    prior_keys = month_keys_through(prior_end, 12)
    cat_recent: Counter = Counter()
    cat_prior: Counter = Counter()
    for a in spending:
        if not a["date"]:
            continue
        key = a["date"].strftime("%Y-%m")
        if key in trend_keys:
            cat_recent[a["cat"]] += a["amount"]
        elif key in prior_keys:
            cat_prior[a["cat"]] += a["amount"]

    recent_total = sum(cat_recent.values()) or 1
    largest_12 = [
        {
            "category": c,
            "amount": money(v),
            "pct": money(v / recent_total * 100),
        }
        for c, v in cat_recent.most_common(6)
    ]

    growth_rows = []
    for cat in set(cat_recent) | set(cat_prior):
        cur = cat_recent.get(cat, 0)
        prev = cat_prior.get(cat, 0)
        if cur < 50:
            continue
        if prev <= 0:
            change = 100.0 if cur > 0 else 0.0
        else:
            change = (cur - prev) / prev * 100
        growth_rows.append(
            {
                "category": cat,
                "amount": money(cur),
                "priorAmount": money(prev),
                "changePct": money(change),
            }
        )
    growth_rows.sort(key=lambda r: -r["changePct"])
    fastest_growing = growth_rows[:5]

    savings = []
    grok_monthly = next((r["monthly"] for r in active_subs if r["vendor"] == "Grok"), 0)
    chatgpt_monthly = next((r["monthly"] for r in active_subs if r["vendor"] == "ChatGPT"), 0)
    if grok_monthly > 0 and chatgpt_monthly > 0:
        savings.append(
            {
                "id": "grok-chatgpt",
                "label": "Grok vs ChatGPT overlap",
                "estimateMonthly": money(min(grok_monthly, chatgpt_monthly)),
                "detail": "Both billed — pick one primary AI assistant",
                "status": "Review",
            }
        )
    churn_tools = [r for r in active_subs if r["vendor"] in {"GenSpark", "Youmind", "Abacus", "Creative Fabrica", "KITTL"} and r["monthly"] < 50]
    if churn_tools:
        churn_total = sum(r["monthly"] for r in churn_tools)
        savings.append(
            {
                "id": "ai-churn",
                "label": "AI trial tools",
                "estimateMonthly": money(churn_total),
                "detail": f"{len(churn_tools)} smaller AI tools — consolidate or cancel",
                "status": "Watch",
            }
        )
    stream_monthly = sum(
        r["monthly"]
        for r in active_subs
        if r["vendor"] in {"YouTube TV", "YouTube Premium", "Netflix", "HBO Max", "Paramount+", "Hulu", "Peacock", "Max"}
    )
    if stream_monthly > 60:
        savings.append(
            {
                "id": "streaming",
                "label": "Streaming stack",
                "estimateMonthly": money(stream_monthly * 0.25),
                "detail": f"~${stream_monthly:,.0f}/mo across streaming — trim unused services",
                "status": "Watch",
            }
        )
    amazon_recent = cat_recent.get("Amazon", 0) + cat_recent.get("Personal", 0)
    if amazon_recent > 3000:
        savings.append(
            {
                "id": "amazon-discipline",
                "label": "Amazon + personal shopping",
                "estimateMonthly": money(amazon_recent / 12 * 0.15),
                "detail": "Largest discretionary bucket — 15% discipline target",
                "status": "Watch",
            }
        )

    # Cash flow — last 12 calendar months from nebat
    month_in = Counter()
    month_out = Counter()
    for n in nebat:
        if not n["date"]:
            continue
        key = n["date"].strftime("%Y-%m")
        if n["deposit"] > 0 and (n["account"] or "") not in {"Open"}:
            if n["account"] not in ONE_TIME_INCOME:
                month_in[key] += n["deposit"]
        if n["withdrawal"] > 0 and n["account"] not in TRANSFER_WITHDRAWALS:
            month_out[key] += n["withdrawal"]

    # also add apple spend by month as "out" overlay (card purchases)
    apple_month = Counter()
    for a in spending:
        if a["date"]:
            apple_month[a["date"].strftime("%Y-%m")] += a["amount"]

    all_months = sorted(set(month_in) | set(month_out) | set(apple_month))[-12:]
    cash_months = []
    for key in all_months:
        inn = month_in.get(key, 0)
        card = apple_month.get(key, 0)
        out = month_out.get(key, 0) + card
        cash_months.append(
            {
                "month": datetime.strptime(key, "%Y-%m").strftime("%b %y"),
                "in": money(inn),
                "out": money(out),
                "net": money(inn - out),
            }
        )

    recent = [m for m in cash_months if m["in"] > 0][-6:] or cash_months[-6:]
    avg_in = sum(m["in"] for m in recent) / max(len(recent), 1)
    avg_out = sum(m["out"] for m in recent) / max(len(recent), 1)
    monthly_net = avg_in - avg_out
    nebat_balance = next((n["balance"] for n in reversed(nebat) if n["balance"] is not None), 0.0) or 0.0

    # Top categories
    cat_totals = Counter()
    for a in spending:
        cat_totals[a["cat"]] += a["amount"]
    top_categories = [
        {"category": c, "amount": money(v), "pct": money(v / apple_lifetime * 100)}
        for c, v in cat_totals.most_common(10)
    ]

    # Review queue
    reviews = []
    if ai_flags:
        for i, flag in enumerate(ai_flags):
            reviews.append({"id": f"ai-{i}", "label": "AI subscriptions", "detail": flag, "status": "Review"})
    if sub_rows and sub_rows[0]["monthly"] > 40:
        reviews.append(
            {
                "id": "youtube-tv",
                "label": "Streaming stack",
                "detail": "YouTube TV + Premium + Netflix — confirm what you still watch",
                "status": "Watch",
            }
        )
    if amazon_monthly > 500:
        reviews.append(
            {
                "id": "amazon",
                "label": "Amazon volume",
                "detail": f"~${amazon_monthly:,.0f}/mo on Amazon — largest discretionary line",
                "status": "Watch",
            }
        )
    if retro_ops_monthly >= 500:
        reviews.append(
            {
                "id": "retro-ops",
                "label": "Retroverse ops",
                "detail": f"Ops run rate ~${retro_ops_monthly:,.0f}/mo — review gear vs cloud",
                "status": health_retro(retro_ops_monthly),
            }
        )
    if ai_run_rate >= 1800:
        reviews.append(
            {
                "id": "local-gpu",
                "label": "Local AI hardware",
                "detail": f"Cloud AI ~${ai_run_rate:,.0f}/yr — worth modeling a local GPU break-even",
                "status": "Watch",
            }
        )

    snapshot = {
        "generatedAt": datetime.now().isoformat(),
        "workbookPath": str(workbook),
        "workbookMtime": mtime,
        "periodLabel": f"YTD {current_year}",
        "dataThrough": data_through_iso,
        "income": {
            "monthlyEstimate": money(income_monthly),
            "ytd": money(income_ytd),
            "status": "Healthy",
            "sources": [
                {"label": k, "amount": money(v)}
                for k, v in income_by_acct.most_common()
                if k not in ONE_TIME_INCOME
            ][:4],
        },
        "requiredBills": {
            "monthlyEstimate": money(bills_monthly),
            "status": "Healthy" if bills_monthly < income_monthly * 0.45 else "Watch",
            "lines": [{"label": k, "amount": money(v / months_span)} for k, v in bill_lines.most_common()[:8]],
        },
        "appleCard": {
            "lifetime": money(apple_lifetime),
            "ytd": money(apple_ytd),
            "monthlyAvg": money(apple_monthly),
            "status": "Watch" if apple_ytd / max(ytd_months, 1) > income_monthly * 0.9 else "Healthy",
        },
        "amazon": {
            "lifetime": money(amazon_lifetime),
            "ytd": money(amazon_ytd),
            "monthlyAvg": money(amazon_monthly),
            "status": "Watch",
            "note": "Itemized detail only — most Amazon spend is on Apple Card already",
        },
        "ai": {
            "monthlyAvg": money(ai_monthly_ytd),
            "ytd": money(ai_ytd),
            "runRateAnnual": money(ai_run_rate),
            "lifetime": money(ai_lifetime),
            "pctOfAppleSpend": money(ai_ytd / apple_ytd * 100 if apple_ytd else 0),
            "pctChangeVsPriorYear": money(pct_change),
            "status": health_ai(ai_monthly_ytd, pct_change),
            "tools": ai_tools[:8],
            "flags": ai_flags,
            "trend": {
                "months": ai_trend_months,
                "monthlyAvg12": ai_trend_avg,
                "annualProjection": ai_trend_annual,
            },
        },
        "retroverse": {
            "opsMonthly": money(retro_ops_monthly),
            "opsYtd": money(retro_ops_ytd),
            "gearYtd": money(retro_gear_ytd),
            "hostingMonthly": money(hosting_monthly),
            "annualEstimate": money(retro_ops_monthly * 12),
            "status": health_retro(retro_ops_monthly),
            "lines": [{"label": k, "amount": money(v / months_span)} for k, v in retro_ops.most_common()[:6]],
            "trend": {
                "months": retro_trend_months,
                "monthlyAvg12": retro_trend_avg,
                "annualProjection": retro_trend_annual,
            },
        },
        "subscriptions": {
            "monthlyTotal": money(sum(r["monthly"] for r in sub_rows)),
            "annualTotal": money(sum(r["monthly"] for r in sub_rows) * 12),
            "status": "Review" if any(r["status"] == "Review" for r in top_subs) else "Watch",
            "active": active_subs,
            "top": top_subs,
            "moreCount": len(more),
            "moreMonthly": money(more_monthly),
        },
        "cashFlow": {
            "status": health_cashflow(monthly_net),
            "monthlyNet": money(monthly_net),
            "nebatBalance": money(nebat_balance),
            "months": cash_months,
        },
        "topCategories": top_categories,
        "opportunity": {
            "largest": largest_12,
            "fastestGrowing": fastest_growing,
            "potentialSavings": savings,
        },
        "reviewNeeded": reviews,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(snapshot, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
