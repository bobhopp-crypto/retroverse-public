#!/usr/bin/env python3
"""Extract unique APPLE sheet column E (Category) values from workbook. stdout = JSON."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

DEFAULT = Path("/Users/bobhopp/FINANCIAL/2021-2026 Financial Workbook.xlsx")


def slugify(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return s[:80] or "account"


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT
    wb = load_workbook(path, read_only=True, data_only=True)
    counts: dict[str, int] = {}

    for i, row in enumerate(wb["APPLE"].iter_rows(values_only=True)):
        if i == 0:
            continue
        raw = row[4]
        if raw is None:
            continue
        name = str(raw).strip()
        if not name:
            continue
        counts[name] = counts.get(name, 0) + 1

    wb.close()

    seen_slugs: dict[str, int] = {}
    accounts = []
    for name, txn_count in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
        base = slugify(name)
        n = seen_slugs.get(base, 0)
        seen_slugs[base] = n + 1
        slug = base if n == 0 else f"{base}-{n + 1}"
        accounts.append(
            {
                "name": name,
                "slug": slug,
                "workbookTxnCount": txn_count,
                "active": name not in {"Payment", "Credit", "Debit"},
            }
        )

    json.dump({"accounts": accounts, "total": len(accounts)}, sys.stdout, indent=2)


if __name__ == "__main__":
    main()
