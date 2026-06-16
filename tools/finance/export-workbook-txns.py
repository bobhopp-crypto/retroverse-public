#!/usr/bin/env python3
"""Export workbook transactions as JSON (read-only). stdout only."""

from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

DEFAULT = Path("/Users/bobhopp/FINANCIAL/2021-2026 Financial Workbook.xlsx")
EXCLUDE = {"Payment", "Credit", "Debit"}

CATEGORY_SLUG = {
    "AI - ChatGPT": "ai-chatgpt",
    "Software - ChatGPT": "ai-chatgpt",
    "AI - Cursor": "ai-cursor",
    "AI - Grok": "ai-grok",
    "SUB - Grok": "ai-grok",
    "AI - Gemini": "ai-gemini",
    "AI - Creative Fabrica": "ai-creative-fabrica",
    "3D Printing": "retro-3d-printing",
    "Engraving": "retro-printing",
    "Web - Cloudflare": "retro-hosting",
    "Web - Neon.tech": "retro-hosting",
    "Web - Netlify": "retro-hosting",
    "Home": "home",
    "Grocery": "grocery",
    "Restaurants": "restaurants",
    "Personal": "personal",
    "Amazon": "amazon",
}


def parse_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []

    for i, row in enumerate(wb["APPLE"].iter_rows(values_only=True)):
        if i == 0:
            continue
        amt = row[6]
        if amt is None or float(amt) <= 0:
            continue
        cat = str(row[4] or "")
        if cat in EXCLUDE:
            continue
        dt = parse_date(row[0])
        if not dt:
            continue
        merchant = str(row[3] or "Apple Card")
        desc = str(row[2] or merchant)
        amount = float(amt)
        out.append(
            {
                "source": "apple_card",
                "transactionDate": dt,
                "merchant": merchant,
                "description": desc,
                "amount": amount,
                "accountName": cat or None,
                "subcategory": cat or None,
                "dedupeKey": f"apple_card|{dt}|{amount:.2f}|{merchant}|{desc[:80]}",
            }
        )

    for i, row in enumerate(wb["AMAZON"].iter_rows(values_only=True)):
        if i == 0:
            continue
        try:
            amount = float(row[4])
        except (TypeError, ValueError):
            continue
        if amount <= 0:
            continue
        dt = parse_date(row[0])
        if not dt:
            continue
        title = str(row[3] or "Amazon")
        acct = str(row[2] or "Personal")
        out.append(
            {
                "source": "amazon",
                "transactionDate": dt,
                "merchant": "Amazon",
                "description": title,
                "amount": amount,
                "categorySlug": "retro-equipment" if "inventory" in acct.lower() else "amazon",
                "subcategory": acct,
                "dedupeKey": f"amazon|{dt}|{amount:.2f}|Amazon|{title[:80]}",
            }
        )

    wb.close()
    json.dump(out, sys.stdout)


if __name__ == "__main__":
    main()
